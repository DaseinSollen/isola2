#---- Programma Barcode per Isola Ecologica ----#

import os
import serial
import signal
import sys
import datetime
import openpyxl
from openpyxl.styles import Alignment

def signal_handler(sig, frame):
    print('Il programma è stato terminato manualmente')
    ser.close()
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)

# porta seriale per S.O. Windows
ser = serial.Serial('COM7', 9600, timeout=1)

# porta per S.O. MacOS // non fissa, da ricontrollare ogni volta
# comando Mac nel terminale || ls /dev/tty.*
#ser = serial.Serial('/dev/tty.usbmodem1444201', 9600, timeout=1)

# Genera il nome del file Excel con il formato gg-mm-aaaa.xlsx
today = datetime.datetime.today().strftime('%d-%m-%Y')
filename = f"{today}.xlsx"

if os.path.isfile(filename):
    # Carica la cartella di lavoro se già esiste
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    # Trova l'ID massimo presente per tracciare l'ultima scansione
    max_id = 0
    for cell in worksheet['A'][1:]:
        if cell.value is not None:
            max_id = max(max_id, int(cell.value))

    # Inizia a scrivere i nuovi dati dopo l'ultima scansione esistente
    scan_id = max_id + 1

    current_time = datetime.datetime.now().strftime('%H:%M:%S')

else:
    # Crea una nuova cartella di lavoro se non esiste
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Aggiunge l'intestazione per le 3 colonne
    worksheet["A1"] = "ID"
    worksheet["B1"] = "Orario"
    worksheet["C1"] = "Scansione"

    # Inizia a registrare nuove scansioni a partire dall'ID 1
    scan_id = 1

disconnected = False

while True:
    try:
        line = ser.readline().decode('utf-8').rstrip()
        if line:
            # Ottiene la data odierna
            current_time = datetime.datetime.now().strftime("%H:%M:%S")

            # Aggiunge la scansione al file Excel gg-mm-aaaa.xslx
            worksheet.cell(row=scan_id+1, column=1, value=scan_id).alignment = Alignment(horizontal='center')
            worksheet.cell(row=scan_id+1, column=2, value=current_time)
            worksheet.cell(row=scan_id+1, column=3, value=line)
            
            # Incrementa l'ID di 1
            scan_id += 1
            
            # Salva il file Excel
            workbook.save(filename)
            
            # Stampa la scansione nel terminale 
            print(f"Barcode letto: {line} - {current_time}")
            
            disconnected = False
            
    except KeyboardInterrupt:
        break
        
    except serial.SerialException:
        if not disconnected:
            print("Il lettore è stato sconnesso, ricollegarlo per procedere")
            disconnected = True
            
            # Aspetta che il lettore venga ricollegato
            while True:
                try:
                    # porta seriale per S.O. Windows
                    ser = serial.Serial('COM7', 9600, timeout=1)

                    # porta per S.O. MacOS
                    #ser = serial.Serial('/dev/tty.usbmodem1444201', 9600, timeout=1)
                    print("Il lettore è stato ricollegato correttamente")
                    disconnected = False
                    break
                except serial.SerialException:
                    pass

    except KeyboardInterrupt:
        break

ser.close()
